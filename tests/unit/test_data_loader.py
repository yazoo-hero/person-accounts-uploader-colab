"""Unit tests for data loading functionality."""
import pytest
import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from src.core.data_loader import DataLoader, WorkdayDataLoader, CalabrioDataLoader
from src.utils.exceptions import DataLoadError


class TestDataLoader:
    """Test cases for DataLoader class."""
    
    def test_load_excel_files_success(self, tmp_path):
        """Test successful Excel file loading."""
        # Create test Excel file
        excel_file = tmp_path / "test_data.xlsx"
        test_df = pd.DataFrame({
            "col1": [1, 2, 3],
            "col2": ["a", "b", "c"]
        })
        test_df.to_excel(excel_file, index=False)
        
        # Load and verify
        result = DataLoader.load_excel_files(tmp_path)
        
        assert len(result) == 3
        assert "col1" in result.columns
        assert "col2" in result.columns
    
    def test_load_excel_files_with_skiprows(self, tmp_path):
        """Test Excel file loading with skiprows."""
        # Create test Excel file with extra rows at the top
        excel_file = tmp_path / "test_data.xlsx"
        
        # Create a workbook with manual data entry
        wb = Workbook()
        ws = wb.active
        
        # Add header rows to skip
        ws.append(["Header Row 1", "Header Row 1"])
        ws.append(["Header Row 2", "Header Row 2"])
        
        # Add actual headers
        ws.append(["col1", "col2"])
        
        # Add data rows
        ws.append([1, "a"])
        ws.append([2, "b"])
        ws.append([3, "c"])
        
        wb.save(excel_file)
        
        # Load skipping first 2 rows
        result = DataLoader.load_excel_files(tmp_path, skiprows=2)
        
        assert len(result) == 3
        assert list(result.columns) == ["col1", "col2"]
    
    def test_load_excel_files_directory_not_exists(self):
        """Test loading from non-existent directory."""
        with pytest.raises(DataLoadError) as exc_info:
            DataLoader.load_excel_files(Path("/nonexistent/directory"))
        
        assert "Directory does not exist" in str(exc_info.value)
    
    def test_load_excel_files_no_files(self, tmp_path):
        """Test loading when no Excel files exist."""
        result = DataLoader.load_excel_files(tmp_path)
        
        assert result.empty
    
    def test_load_excel_files_latest_file(self, tmp_path):
        """Test that the latest file is loaded when multiple exist."""
        # Create multiple Excel files
        for i in range(3):
            excel_file = tmp_path / f"test_data_{i}.xlsx"
            test_df = pd.DataFrame({"value": [i]})
            test_df.to_excel(excel_file, index=False)
            # Ensure different timestamps
            import time
            time.sleep(0.01)
        
        # Load and verify latest file is used
        result = DataLoader.load_excel_files(tmp_path)
        
        assert result["value"].iloc[0] == 2  # Latest file should have value 2
    
    def test_load_json_file_success(self, tmp_path):
        """Test successful JSON file loading."""
        # Create test JSON file
        json_file = tmp_path / "test_data.json"
        test_data = [
            {"id": 1, "name": "Alice"},
            {"id": 2, "name": "Bob"}
        ]
        json_file.write_text(json.dumps(test_data))
        
        # Load and verify
        result = DataLoader.load_json_file(json_file)
        
        assert len(result) == 2
        assert "id" in result.columns
        assert "name" in result.columns
    
    def test_load_json_file_not_exists(self):
        """Test loading non-existent JSON file."""
        with pytest.raises(DataLoadError) as exc_info:
            DataLoader.load_json_file(Path("/nonexistent/file.json"))
        
        assert "JSON file does not exist" in str(exc_info.value)
    
    def test_load_json_file_invalid_json(self, tmp_path):
        """Test loading invalid JSON file."""
        json_file = tmp_path / "invalid.json"
        json_file.write_text("{ invalid json }")
        
        with pytest.raises(DataLoadError) as exc_info:
            DataLoader.load_json_file(json_file)
        
        assert "Invalid JSON" in str(exc_info.value)


class TestWorkdayDataLoader:
    """Test cases for WorkdayDataLoader class."""
    
    def test_load_all_data(self, tmp_path):
        """Test loading all Workday data."""
        # Create directory structure
        person_accounts_dir = tmp_path / "person_accounts"
        people_dir = tmp_path / "people"
        used_entries_dir = tmp_path / "used_entries"
        
        person_accounts_dir.mkdir()
        people_dir.mkdir()
        used_entries_dir.mkdir()
        
        # Create test data files with header rows to match actual files
        # Person accounts (skiprows=6)
        wb = Workbook()
        ws = wb.active
        for _ in range(6):  # Add 6 header rows
            ws.append(["Header"] * 2)
        ws.append(["WiserId", "Beginning Year Balance"])  # Column headers
        ws.append(["123", 10])
        ws.append(["456", 20])
        wb.save(person_accounts_dir / "data.xlsx")
        
        # People data (skiprows=2)
        wb = Workbook()
        ws = wb.active
        for _ in range(2):  # Add 2 header rows
            ws.append(["Header"] * 2)
        ws.append(["Latest Headcount Wiser ID", "Latest Headcount Hire Date"])
        ws.append(["123", "2020-01-01"])
        ws.append(["456", "2021-01-01"])
        wb.save(people_dir / "data.xlsx")
        
        # Used entries (skiprows=6)
        wb = Workbook()
        ws = wb.active
        for _ in range(6):  # Add 6 header rows
            ws.append(["Header"] * 2)
        ws.append(["WiserId", "Units Used"])
        ws.append(["123", 5])
        wb.save(used_entries_dir / "data.xlsx")
        
        # Load data
        loader = WorkdayDataLoader(person_accounts_dir, people_dir, used_entries_dir)
        workday_df, people_df, used_entries_df = loader.load_all_data()
        
        # Verify
        assert len(workday_df) == 2
        assert len(people_df) == 2
        assert len(used_entries_df) == 1
        assert workday_df["WiserId"].dtype == "object"  # String type
        assert "WiserId" in people_df.columns  # Column renamed
        assert pd.api.types.is_datetime64_any_dtype(people_df["Latest Headcount Hire Date"])
    
    def test_load_with_missing_directories(self, tmp_path):
        """Test loading with missing directories."""
        loader = WorkdayDataLoader(
            tmp_path / "missing1",
            tmp_path / "missing2",
            tmp_path / "missing3"
        )
        workday_df, people_df, used_entries_df = loader.load_all_data()
        
        # Should return empty DataFrames
        assert workday_df.empty
        assert people_df.empty
        assert used_entries_df.empty


class TestCalabrioDataLoader:
    """Test cases for CalabrioDataLoader class."""
    
    def test_load_all_data(self, tmp_path):
        """Test loading all Calabrio data."""
        # Create test JSON files
        account_data = [
            {
                "EmploymentNumber": "123",
                "Accrued": "10.5",
                "BalanceIn": 20
            },
            {
                "EmploymentNumber": "456",
                "Accrued": "invalid",  # Test error handling
                "BalanceIn": 30
            }
        ]
        person_data = [
            {
                "EmploymentNumber": "123",
                "EmploymentStartDate": "2020-01-15T00:00:00"
            },
            {
                "EmploymentNumber": "456",
                "EmploymentStartDate": "2021-06-01T00:00:00"
            }
        ]
        
        account_file = tmp_path / "account_data.json"
        person_file = tmp_path / "person_data.json"
        
        account_file.write_text(json.dumps(account_data))
        person_file.write_text(json.dumps(person_data))
        
        # Load data
        loader = CalabrioDataLoader(account_file, person_file)
        calabrio_df, person_df = loader.load_all_data()
        
        # Verify
        assert len(calabrio_df) == 2
        assert len(person_df) == 2
        assert calabrio_df["EmploymentNumber"].dtype == "object"  # String type
        assert calabrio_df.loc[0, "Accrued"] == 10.5
        assert calabrio_df.loc[1, "Accrued"] == 0  # Invalid value becomes 0
        assert pd.api.types.is_datetime64_any_dtype(person_df["EmploymentStartDate"])
    
    def test_load_with_missing_files(self, tmp_path):
        """Test loading with missing files."""
        loader = CalabrioDataLoader(
            tmp_path / "missing1.json",
            tmp_path / "missing2.json"
        )
        calabrio_df, person_df = loader.load_all_data()
        
        # Should return empty DataFrames
        assert calabrio_df.empty
        assert person_df.empty